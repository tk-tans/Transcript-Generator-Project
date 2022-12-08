def range_generate(range_roll):
    import os
    from typing import Collection
    from fpdf.fpdf import TitleStyle
    import openpyxl
    from fpdf import FPDF
    import csv
    from collections import defaultdict
    from datetime import datetime
    import pytz
    import shutil
    not_present = []
    department={"CS":"Computer Science and Engineering  ","EE":"Electrical and Electronics Engineering ","ME":"Mechanical Engineering ","CE":"Civil Engineering ","MM":"Metallurgy and Materials Engineering "}

    programme={"01":"Bachelor of Technology","11":"Masters in Technology"}

    grades_to_points={
        "AA" : 10,"AB" : 9,"BB" : 8,"BC" : 7,"CC" : 6,"CD" : 5,"DD" : 4,"F"  : 0,"I"  : 0 }
    # grades={"AA*":"AA","AA":"AA","AB":"AB","AB*":"AB","BB":"BB","BB*":"BB","BC":"BC","BC*":"BC","CC":"CC","CC*":"CC","CD":"CD","CD*":"CD","DD*":"DD","DD":"DD","F*":"F","F":"F","I*":"I","I":"I"}

    subject_name ={}
    subject_ltp={}
    roll_and_name={}

    def cpi_genertor(sem_no,record):
        i=2
        total_credit=record["1"][0]
        cpi=record["1"][1]
        for i in range(2,int(sem_no)+1):
            if i==9:
                continue
            current_sem_credit=record[str(i)][0]
            current_sem_cpi=record[str(i)][1]
            total=float(current_sem_cpi*current_sem_credit)+float(total_credit*cpi)
            total_credit+=record[str(i)][0]
            cpi1=total/(total_credit )
            cpi=round(cpi1,2)
        return cpi
        
            
        print("hello")
    def rollno_and_name_extractor():
        location="uploads\\names-roll.csv"
        with open(location,"r")as f:
            reader=csv.DictReader(f)
            for row in reader:
                dct=dict(row)
                roll_and_name[dct["Roll"]]=dct["Name"]


    def credit_cleared(data):
        cleared_dict = {}
        for sem in data:
            cleared_credit=0
            for record in data[sem]:
                if record["Grade"] =="F*" or record["Grade"] =="F" or record["Sub_Type"]=="Backlog":
                    continue
                else:
                    cleared_credit+=int(record["Credit"])

            cleared_dict[sem]=cleared_credit
        
        # print(cleared_dict)
        # //exit()
        return cleared_dict

    def spi_calculator( record):
        credit=[]
        grades=[]
        for row in record :
            credit.append(row[0])
            grades.append(row[1])
        
        credit=[int(score) for score in credit]
        grades=[int(grades_to_points[(j.strip('*')).strip()])for j in grades]
        sem_credit=sum(credit)
        score_of_sem=[int(i*j)for i,j in zip(credit,grades)]
        spi=0
        if (int(sem_credit)):
            spi=float(sum(score_of_sem))/float(sem_credit)
        spi=round(spi,2)
        return sem_credit,spi
    def subject_information():
        location="uploads\\subjects_master.csv"
        with open(location,"r")as f:
            reader=csv.DictReader(f)
            for row in reader:
                dct=dict(row)
                subject_ltp[dct["subno"]]=dct["ltp"]
                subject_name[dct["subno"]]=dct["subname"]
    def create_table(table_data, title='', data_size = 10, title_size=12, align_data='C', align_header='C', cell_width=20, x_start='x_default',emphasize_data=[], emphasize_style=None, emphasize_color=(0,0,0)):
        default_style = workpdf.font_style
        if emphasize_style == None:
            emphasize_style = default_style
        # Get Width of Columns
        def get_col_widths():
            col_width = cell_width
            if isinstance(cell_width, list):
                new_list=[]
                for i in cell_width:
                    new_list.append(int(i))
                col_width=new_list
                    
            # TODO: convert all items in list to int        
            return col_width  
        header = table_data[0]
        data = table_data[1:]
        line_height = 4.5
        col_width = get_col_widths()
        workpdf.set_font("Helvetica","BU",12)
        # Get starting position of x
        # Determin width of table to get x starting point for centred table
        if x_start == 'C':
            table_width = 0
            if isinstance(col_width, list):
                for width in col_width:
                    table_width += width
            else: # need to multiply cell width by number of cells to get table width 
                table_width = col_width * len(table_data[0])
            # Get x start by subtracting table width from workpdf width and divide by 2 (margins)
            margin_width = workpdf.w - table_width
            # TODO: Check if table_width is larger than workpdf width

            center_table = margin_width / 2 # only want width of left margin not both
            x_start = center_table
            workpdf.set_x(x_start)
        elif isinstance(x_start, int):
            workpdf.set_x(x_start)
        elif x_start == 'x_default':
            x_start = workpdf.set_x(workpdf.l_margin)
        # TABLE CREATION #
        # add title
        if title != '':
            workpdf.set_font("Helvetica","BU",12)
            workpdf.multi_cell(0, 3, title, border=0, align='j', ln=3, max_line_height=workpdf.font_size)
            workpdf.ln(line_height) # move cursor back to the left margin
        # add header
        workpdf.set_font("Helvetica","B",10)
        y1 = workpdf.get_y()
        if x_start:
            x_left = float(x_start)
        else:
            x_left = float(workpdf.get_x())
        x_right = workpdf.epw + x_left
    
        if isinstance(col_width, list):
            if x_start:
                workpdf.set_x(x_start)
            for i in range(len(header)):
                datum = header[i]
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(col_width[i], line_height, datum, border=1, align=align_header, ln=3, max_line_height=workpdf.font_size)
                x_right = workpdf.get_x()
            workpdf.ln(line_height) # move cursor back to the left margin
            y2 = workpdf.get_y()
            workpdf.line(x_left,y1,x_right,y1)
            workpdf.line(x_left,y2,x_right,y2)

            workpdf.set_font("Helvetica","",8)
            for i in range(len(data)):
                if x_start:
                    workpdf.set_x(x_start)
                row = data[i]
                for i in range(len(row)):
                    datum = row[i]
                    if not isinstance(datum, str):
                        datum = str(datum)
                    adjusted_col_width = col_width[i]
                    if datum in emphasize_data:
                        workpdf.set_text_color(*emphasize_color)
                        workpdf.set_font(style=emphasize_style)
                        workpdf.multi_cell(adjusted_col_width, line_height, datum, border=1, align=align_data, ln=3, max_line_height=workpdf.font_size)
                        workpdf.set_text_color(0,0,0)
                        workpdf.set_font(style=default_style)
                    else:
                        workpdf.multi_cell(adjusted_col_width, line_height, datum, border=1, align=align_data, ln=3, max_line_height=workpdf.font_size) 
                        
                        # ln = 3 - move cursor to right with same vertical offset # this uses an object named workpdf
            
                workpdf.ln(line_height) # move cursor back to the left margin
            # if footer_string !='':
            #     workpdf.set_font("Helvetica","B",12)
            #     workpdf.multi_cell(0, 3, footer_string, border=0, align='j', ln=3, max_line_height=workpdf.font_size)
            #     workpdf.ln(line_height) # move cursor back to the left margin
        y3 = workpdf.get_y()
        workpdf.line(x_left,y3,x_right,y3)
    
    if os.path.exists(".\\outputs"):
        shutil.rmtree(".\outputs")
    if os.path.exists(".\Transcipts.zip"):
        os.remove(".\Transcipts.zip")
    rollno_and_name_extractor()
    subject_information()
    for i in range_roll:
        if i not in roll_and_name.keys():
            not_present.append(i)
    for roll in roll_and_name.keys():
        if roll.upper() not in range_roll:
            continue
        d = defaultdict(list)
        data=[]
        
        header_list=["Sub.Code","Subject Name","L-T-P","CRD","GRD"]
        data.append(header_list)
        location="uploads\\grades.csv"
        with open(location,"r")as f:
            reader =csv.DictReader(f)
            for row in reader :
                dct=dict(row)
                sem_no =dct["Sem"]

                if roll == dct["Roll"]:
                    d[sem_no].append(dct)
        #print(d)
        # print(data)
        cleared_dict = {}
        cleared_dict = credit_cleared(d)

        workpdf=FPDF(orientation="L",unit="mm",format="A3")
        workpdf.add_page()
        # workpdf=FPDF(orientation="L",unit="mm",format="A3")
        workpdf.rect(10,5,400,288)
        workpdf.rect(10,5,74,35)
        workpdf.rect(336,5,74,35)
        workpdf.rect(84,5,252,35)
        workpdf.rect(10,40,400,75)
        workpdf.rect(10,115,400,60)
        workpdf.rect(10,175,400,58)
        print(type(workpdf.epw))
        # workpdf.add_font("Lucida Fax Demibold Italic","",r"C:\Windows\Fonts\Lucida Fax\LFAXDI.TTF")
        workpdf.image("images\iitplogo.png",25,8,40,24,'PNG')
        workpdf.set_xy(19,22)
        workpdf.set_font("Helvetica",'BU',14)
        workpdf.cell(35,30,"INTERIM TRANSCRIPT")
        workpdf.image("images\iitplogo.png",356,8,40,24,'PNG')
        workpdf.set_xy(349,22)
        workpdf.set_font("Helvetica",'BU',14)
        workpdf.cell(324,30,"INTERIM TRANSCRIPT")
        workpdf.image("images\iitphindilogo.png",90,6,240,16,'PNG')
        workpdf.set_font("Helvetica",'B',26)
        workpdf.ln(10)
        workpdf.set_xy(120,16)
        workpdf.cell(100,22,"Indian Institute of Technology Patna")
        workpdf.set_xy(180,22)
        workpdf.set_font("Helvetica",'B',18)
        workpdf.cell(120,30,"Transcript")
        workpdf.rect(124,42,210,12)
        workpdf.set_font("Helvetica",'B',14)
        workpdf.set_xy(125,23)
        workpdf.cell(24,44,"Roll No:")
        workpdf.rect(149,43,25,5)
        workpdf.set_font("Helvetica",'',12)
        workpdf.cell(152,45,f"{roll}")
        workpdf.set_xy(180,23)
        workpdf.set_font("Helvetica",'B',12)
        workpdf.cell(93,44,"Name:")
        workpdf.rect(197,43,55,5)
        workpdf.set_font("Helvetica",'',12)
        workpdf.set_xy(200,23)
        workpdf.cell(101,45,f"{roll_and_name[roll]}")
        workpdf.set_xy(260,23)
        workpdf.set_font("Helvetica",'B',14)
        workpdf.cell(150,44,"Year of Admission:")
        workpdf.rect(310,43,20,5)
        year_of_admission="20"+str(roll[0:2])
        workpdf.set_font("Helvetica",'',12)
        workpdf.set_xy(312,23)
        workpdf.cell(273,45,f"{year_of_admission}")
        workpdf.set_xy(125,34)
        workpdf.set_font("Helvetica",'B',14)
        workpdf.cell(24,34,"Programme:")
        workpdf.set_xy(155,34)
        workpdf.set_font("Helvetica",'',14)
        workpdf.cell(160,34,f"{programme[roll[2:4]]}")
        workpdf.set_font("Helvetica",'B',14)
        workpdf.set_xy(230,34)
        workpdf.cell(230,34,"Course:")
        workpdf.set_xy(250,34)
        workpdf.set_font("Helvetica",'',14)
        workpdf.cell(248,34,f"{department[roll[4:6]]}")
    
        count=0
        colwidth1=30
        colwidth2=70
        colwidth=20
        record_of_credit_and_spi=defaultdict(list)
        for sem_data in d:
            # print(sem_data)
            record=defaultdict(list)
            
            while(count != 0):
                data.pop(count)
                count-=1   
            for row in d[sem_data]:
                list_of_subject_performance=[row["SubCode"],subject_name[row["SubCode"]],subject_ltp[row["SubCode"]],row["Credit"],row["Grade"]]
                data.append(list_of_subject_performance)
                count+=1
                temp=[row["Credit"],row["Grade"]]
                record[sem_data].append(temp)
            # workpdf.set_font("Times",size=10)
            # col_width=workpdf.epw /3
            # line_height=workpdf.font_size*1
            # workpdf.set_font("Helvetica","BU",8)
            # workpdf.multi_cell(0, line_height,f"Semester {sem_data}", border=0, align='j', ln=3, max_line_height=workpdf.font_size)
            # workpdf.set_xy(30,90)
            var=0
            credit,spi=spi_calculator(record[sem_data])
            
            record_of_credit_and_spi[sem_data].append(credit)
            record_of_credit_and_spi[sem_data].append(spi)
            if int(sem_data)==1 :
                workpdf.cell(3)
                workpdf.set_y(55)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=14)
                workpdf.set_y(108)
                workpdf.set_x(20)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
                workpdf.ln(3)
                # print(result_string)
                # workpdf.cell(26,104,f"{ "Credits"}")
            elif int(sem_data) ==4 :
                workpdf.cell(3)
                workpdf.set_y(118)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=14)
                workpdf.set_y(167)
                workpdf.set_x(20)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
            
            elif int(sem_data) ==7:
                workpdf.cell(3)
                workpdf.set_y(178)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=14)
                workpdf.set_y(226)
                workpdf.set_x(20)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
                
            elif int(sem_data) ==2 :
                workpdf.cell(5*20)
                workpdf.set_y(55)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=147)
                workpdf.set_y(108)
                workpdf.set_x(147)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
                
                # workpdf.cell(170 ,104,"Credit Taken:  Credit Cleared:  SPI:   CPI:")
            elif int(sem_data) ==5 :
                workpdf.cell(3)
                workpdf.set_y(118)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=147)
                workpdf.set_y(167)
                workpdf.set_x(147)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
                
            elif int(sem_data) ==8 :
                workpdf.cell(3)
                workpdf.set_y(178)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=147)
                workpdf.set_y(226)
                workpdf.set_x(147)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
            
            elif int(sem_data) ==3:
                workpdf.cell(10*20)
                workpdf.set_y(55)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=282)
                workpdf.set_y(108)
                workpdf.set_x(290)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
            
            elif int(sem_data) ==6 :
                workpdf.cell(3)
                workpdf.set_y(118)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=282)
                workpdf.set_y(167)
                workpdf.set_x(290)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size)
            elif int(sem_data) ==10 :
                workpdf.cell(3)
                workpdf.set_y(178)
                cpi=cpi_genertor(sem_data,record_of_credit_and_spi)
                result_str=f"Credit Taken: {record_of_credit_and_spi[sem_data][0]}  Credits Cleared:{cleared_dict[sem_data]}  SPI:{record_of_credit_and_spi[sem_data][1]}  CPI:{cpi}"
                create_table(table_data=data,title=f"Semester{sem_data}",cell_width=[20,70,15,10,10],x_start=282)
                workpdf.set_y(226)
                workpdf.set_x(290)
                workpdf.set_font("Helvetica","B",10)
                workpdf.multi_cell(115, 5, result_str, border=1, align='j', ln=3, max_line_height=workpdf.font_size) 
        workpdf.set_font("Helvetica","B",16)
        workpdf.set_y(218)
        workpdf.set_x(18)
        workpdf.cell(14,59,"Date Generated: ")
        IST=pytz.timezone("Asia/Kolkata")
        xx=datetime.now(IST)
        # time=f'{xx.strftime("%d %b %Y %H : %M")}
        # print(time)
        workpdf.set_y(218)
        workpdf.set_x(65)
        workpdf.cell(14,59,f'{xx.strftime("%d %b %Y, %H:%M")}',align="")
        workpdf.set_y(218)
        workpdf.set_x(250)
        workpdf.cell(14,59,"Assistant Registrar (Academic) :",align="")
        #signature insreting code 
        path_of_image= "signature_photo\\signature.png"
        if os.path.exists(path_of_image):
            workpdf.image(path_of_image,350,236,40,30)
        #seal inserting code 
        path_of_seal="stamp_photo\\stamp.png"
    
        if os.path.exists(path_of_seal):
            workpdf.image(path_of_seal,170,236,40,40)
        # workpdf.image("images\\assistant register .jpeg",330,275,60,15)

        if os.path.exists(os.path.join(".\outputs",str(roll)+".pdf")):
            os.remove(os.path.join(".\outputs",str(roll)+".pdf"))
            workpdf.output(os.path.join(".\outputs",str(roll)+".pdf"))
        else:
            if os.path.exists(".\outputs"):
                workpdf.output(os.path.join(".\outputs",str(roll)+".pdf"))
            else:
                os.mkdir(".\outputs")
                workpdf.output(os.path.join(".\outputs",str(roll)+".pdf"))

    
    return(not_present)
                # for row in data:
                #     var+=1
                #     workpdf.set_font("Helvetica","",8)
                    # if int(sem_data)% 3 ==1:
                    #     workpdf.cell(3)
                    # elif int(sem_data)% 3 ==2:
                    #     workpdf.cell(5*20)
                    # else :
                    #     workpdf.cell(10*20)
                
                    # workpdf.multi_cell(colwidth1,line_height,row[0],border=1,align="C",max_line_height=workpdf.font_size)
                    
                    # workpdf.multi_cell(colwidth2,line_height,row[1],border=1,align="C",max_line_height=workpdf.font_size)
                    
                    # workpdf.multi_cell(colwidth,line_height,row[2],border=1,align="C",max_line_height=workpdf.font_size)
                
                    # workpdf.multi_cell(colwidth,line_height,row[3],border=1,align="C",max_line_height=workpdf.font_size)
                
                    # workpdf.multi_cell(colwidth,line_height,row[4],border=1,align="C",max_line_height=workpdf.font_size)

                    
            # elif sem_data=="2":
            #     while(count != 0):
            #         data.pop(count)
            #         count-=1   
            #     for row in d[sem_data]:
            #         list_of_subject_performance=[row["SubCode"],subject_name[row["SubCode"]],subject_ltp[row["SubCode"]],row["Credit"],row["Grade"]]
            #         data.append(list_of_subject_performance)
            #         count+=1
            # elif sem_data=="3":
                # while(count != 0):
                #     data.pop(count)
                #     count-=1
                # for row in d[sem_data]:
                #     list_of_subject_performance=[row["SubCode"],subject_name[row["SubCode"]],subject_ltp[row["SubCode"]],row["Credit"],row["Grade"]]
                #     data.append(list_of_subject_performance)
                #     count+=1
            # else :
            #     print("hello")
            #for row in sem_data:
    